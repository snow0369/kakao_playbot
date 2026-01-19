from __future__ import annotations

from typing import Dict, Tuple, List, Optional
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from parse import WeaponInfo, ReplyType  # adjust import paths to your project


def export_root_level_enhance_stats_xlsx(
    out_xlsx: str | Path,
    weapon_root_map: Dict[str, Tuple[WeaponInfo, WeaponInfo]],
    enhance_events: Dict[str, Dict[ReplyType, List[int]]],
    *,
    max_level: Optional[int] = None,
) -> None:
    """
    Create an Excel sheet:
      - Columns: root weapon names (each root is a 4-col group: Keep/Success/Break/Samples)
      - Rows: levels; each level occupies 3 rows:
            (1) weapon name (merged across 4 cols)
            (2) counts: Keep/Success/Break/Samples
            (3) probabilities: Pr_keep/Pr_success/Pr_break/Samples

    Missing data -> '-'
    """

    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    # ---- Build indexes: (root_name, level) -> weapon_name ----
    # weapon_root_map maps weapon_name -> (weapon_info(level), root_info(level=1))
    root_names = set()
    root_level_to_weapon: Dict[Tuple[str, int], str] = {}

    inferred_max_level = 0
    for wname, (wi, root) in weapon_root_map.items():
        if wi.level <= 0:
            continue
        root_names.add(root.name)
        key = (root.name, wi.level)
        if key in root_level_to_weapon and root_level_to_weapon[key] != wname:
            raise ValueError(
                f"Multiple weapon names for same (root, level) {key}: "
                f"{root_level_to_weapon[key]} vs {wname}"
            )
        root_level_to_weapon[key] = wname
        inferred_max_level = max(inferred_max_level, wi.level)

    roots_sorted = sorted(root_names)
    if max_level is None:
        max_level = inferred_max_level if inferred_max_level > 0 else 1

    # ---- Workbook / styles ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Enhance Stats"

    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    subheader_fill = PatternFill("solid", fgColor="D9E1F2")  # light
    bold_white = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="808080")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column plan:
    # A: Level
    # For each root: 4 columns (Keep, Success, Break, Samples)
    LEVEL_COL = 1
    start_col = 2

    # ---- Header rows (2 rows) ----
    ws.cell(row=1, column=LEVEL_COL, value="Level")
    ws.merge_cells(start_row=1, start_column=LEVEL_COL, end_row=2, end_column=LEVEL_COL)
    c = ws.cell(row=1, column=LEVEL_COL)
    c.fill = header_fill
    c.font = bold_white
    c.alignment = align_center
    c.border = border_thin

    for i, root in enumerate(roots_sorted):
        base = start_col + 4 * i
        # Root header merged across 4 cols
        ws.cell(row=1, column=base, value=root)
        ws.merge_cells(start_row=1, start_column=base, end_row=1, end_column=base + 3)
        for col in range(base, base + 4):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = bold_white
            cell.alignment = align_center
            cell.border = border_thin

        # Subheaders
        labels = ["Keep", "Success", "Break", "Samples"]
        for j, lab in enumerate(labels):
            cell = ws.cell(row=2, column=base + j, value=lab)
            cell.fill = subheader_fill
            cell.font = bold
            cell.alignment = align_center
            cell.border = border_thin

    # ---- Body: each level occupies 3 rows ----
    # Layout per level:
    #   Row r:     merged weapon name across 4 cols
    #   Row r+1:   counts
    #   Row r+2:   probabilities (and samples)
    row0 = 3
    for lv in range(1, max_level + 1):
        r = row0 + (lv - 1) * 3

        # Merge the Level cell across 3 rows
        ws.cell(row=r, column=LEVEL_COL, value=lv)
        ws.merge_cells(start_row=r, start_column=LEVEL_COL, end_row=r + 2, end_column=LEVEL_COL)
        lvl_cell = ws.cell(row=r, column=LEVEL_COL)
        lvl_cell.font = bold
        lvl_cell.alignment = align_center
        for rr in range(r, r + 3):
            ws.cell(row=rr, column=LEVEL_COL).border = border_thin

        for i, root in enumerate(roots_sorted):
            base = start_col + 4 * i
            wname = root_level_to_weapon.get((root, lv))

            # Pull counts from enhance_events by weapon name (before name)
            if wname is None:
                weapon_display = "-"
                k = s = b = tot = None
            else:
                weapon_display = wname
                ev = enhance_events.get(wname)
                if not ev:
                    k = s = b = tot = 0
                else:
                    k = len(ev.get(ReplyType.ENHANCE_KEEP, []))
                    s = len(ev.get(ReplyType.ENHANCE_SUCCESS, []))
                    b = len(ev.get(ReplyType.ENHANCE_BREAK, []))
                    tot = k + s + b

            # Row r: weapon name merged across 4 cols
            ws.cell(row=r, column=base, value=weapon_display)
            ws.merge_cells(start_row=r, start_column=base, end_row=r, end_column=base + 3)
            for col in range(base, base + 4):
                cell = ws.cell(row=r, column=col)
                cell.alignment = align_left
                cell.border = border_thin

            # Row r+1: counts
            if wname is None:
                vals_counts = ["-", "-", "-", "-"]
            else:
                vals_counts = [k, s, b, tot]
                # If there is no data at all, show '-'
                if tot == 0:
                    vals_counts = ["-", "-", "-", "-"]

            for j, v in enumerate(vals_counts):
                cell = ws.cell(row=r + 1, column=base + j, value=v)
                cell.alignment = align_center
                cell.border = border_thin

            # Row r+2: probabilities + samples
            if wname is None or tot == 0:
                vals_prob = ["-", "-", "-", "-"]
            else:
                vals_prob = [k / tot, s / tot, b / tot, tot]

            for j, v in enumerate(vals_prob):
                cell = ws.cell(row=r + 2, column=base + j, value=v)
                cell.alignment = align_center
                cell.border = border_thin
                # Format probabilities as percent; keep samples as int
                if j < 3 and isinstance(v, (int, float)):
                    cell.number_format = "0.0%"
                if j == 3 and isinstance(v, (int, float)):
                    cell.number_format = "#,##0"

    # ---- Column widths ----
    ws.column_dimensions[get_column_letter(LEVEL_COL)].width = 8
    for i in range(len(roots_sorted)):
        base = start_col + 4 * i
        ws.column_dimensions[get_column_letter(base)].width = 18  # Keep
        ws.column_dimensions[get_column_letter(base + 1)].width = 18  # Success
        ws.column_dimensions[get_column_letter(base + 2)].width = 18  # Break
        ws.column_dimensions[get_column_letter(base + 3)].width = 14  # Samples

    # ---- Row heights ----
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for lv in range(1, max_level + 1):
        r = row0 + (lv - 1) * 3
        ws.row_dimensions[r].height = 18
        ws.row_dimensions[r + 1].height = 18
        ws.row_dimensions[r + 2].height = 18

    ws.freeze_panes = ws["B3"]

    wb.save(out_xlsx)
