from __future__ import annotations

from pathlib import Path
from typing import Dict, Tuple, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from playbot.types import ReplyType, WeaponInfo, TimestampT  # adjust import paths to your project
from playbot.weaponbook import WeaponBook


def export_enhance_stats_xlsx(
    out_xlsx: str | Path,
    weapon_book: WeaponBook,
    enhance_events: Dict[WeaponInfo, Dict[ReplyType, List[Tuple[int, TimestampT]]]],
    # sell_events: Dict[WeaponInfo, List[Tuple[int, TimestampT]]],
    *,
    max_level: int = 20,
) -> None:
    """
    Create an Excel sheet:
      - Columns: root weapon names (each root is a 4-col group: Keep/Success/Break/Samples)
      - Rows: levels; each level occupies 3 rows:
            (1) weapon name (merged across 4 cols)
            (2) counts: Keep/Success/Break/Samples
            (3) probabilities: Pr_keep/Pr_success/Pr_break/Samples

    Missing data -> '-'
    """
    # ---- index -----
    hid_list = sorted(list(weapon_book.hierarchies.keys()))
    special_hids = weapon_book.special_ids

    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    # ---- Workbook / styles ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Enhance Stats"

    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    subheader_fill = PatternFill("solid", fgColor="D9E1F2")  # light
    bold_white = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="808080")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column plan:
    # A: Level
    # For each root: 4 columns (Keep, Success, Break, Samples)
    LEVEL_COL = 1
    start_col = 2

    # ---- Header rows (2 rows) ----
    ws.cell(row=1, column=LEVEL_COL, value="Level")
    ws.merge_cells(start_row=1, start_column=LEVEL_COL, end_row=2, end_column=LEVEL_COL)
    c = ws.cell(row=1, column=LEVEL_COL)
    c.fill = header_fill
    c.font = bold_white
    c.alignment = align_center
    c.border = border_thin

    for i, hid in enumerate(hid_list):
        base = start_col + 4 * i
        # Root header merged across 4 cols
        ws.cell(row=1, column=base, value=hid)
        ws.merge_cells(start_row=1, start_column=base, end_row=1, end_column=base + 3)
        for col in range(base, base + 4):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = bold_white
            cell.alignment = align_center
            cell.border = border_thin

        # Subheaders
        labels = ["Keep", "Success", "Break", "Samples"]
        for j, lab in enumerate(labels):
            cell = ws.cell(row=2, column=base + j, value=lab)
            cell.fill = subheader_fill
            cell.font = bold
            cell.alignment = align_center
            cell.border = border_thin

    # ---- Body: each level occupies 3 rows ----
    # Layout per level:
    #   Row r:     merged weapon name across 4 cols
    #   Row r+1:   counts
    #   Row r+2:   probabilities (and samples)
    row0 = 3
    for lv in range(1, max_level + 1):
        r = row0 + (lv - 1) * 3

        # Merge the Level cell across 3 rows
        ws.cell(row=r, column=LEVEL_COL, value=lv)
        ws.merge_cells(start_row=r, start_column=LEVEL_COL, end_row=r + 2, end_column=LEVEL_COL)
        lvl_cell = ws.cell(row=r, column=LEVEL_COL)
        lvl_cell.font = bold
        lvl_cell.alignment = align_center
        for rr in range(r, r + 3):
            ws.cell(row=rr, column=LEVEL_COL).border = border_thin

        for i, hid in enumerate(hid_list):
            base = start_col + 4 * i
            wnode = weapon_book.hierarchies[hid]["by_level"].get(lv, None)
            wname = wnode["name"] if wnode is not None else None
            # Pull counts from enhance_events by weapon name (before name)
            if wname is None:
                weapon_display = "-"
                k = s = b = tot = None
            else:
                weapon_display = wname if hid not in special_hids else "[특수] "+wname
                ev = enhance_events.get(WeaponInfo(name=wname, level=lv, id=hid))
                if not ev:
                    k = s = b = tot = 0
                else:
                    k = len(ev.get(ReplyType.ENHANCE_KEEP, []))
                    s = len(ev.get(ReplyType.ENHANCE_SUCCESS, []))
                    b = len(ev.get(ReplyType.ENHANCE_BREAK, []))
                    tot = k + s + b

            # Row r: weapon name merged across 4 cols
            ws.cell(row=r, column=base, value=weapon_display)
            ws.merge_cells(start_row=r, start_column=base, end_row=r, end_column=base + 3)
            for col in range(base, base + 4):
                cell = ws.cell(row=r, column=col)
                cell.alignment = align_left
                cell.border = border_thin

            # Row r+1: counts
            if wname is None:
                vals_counts = ["-", "-", "-", "-"]
            else:
                vals_counts = [k, s, b, tot]
                # If there is no data at all, show '-'
                if tot == 0:
                    vals_counts = ["-", "-", "-", "-"]

            for j, v in enumerate(vals_counts):
                cell = ws.cell(row=r + 1, column=base + j, value=v)
                cell.alignment = align_center
                cell.border = border_thin

            # Row r+2: probabilities + samples
            if wname is None or tot == 0:
                vals_prob = ["-", "-", "-", "-"]
            else:
                vals_prob = [k / tot, s / tot, b / tot, tot]

            for j, v in enumerate(vals_prob):
                cell = ws.cell(row=r + 2, column=base + j, value=v)
                cell.alignment = align_center
                cell.border = border_thin
                # Format probabilities as percent; keep samples as int
                if j < 3 and isinstance(v, (int, float)):
                    cell.number_format = "0.0%"
                if j == 3 and isinstance(v, (int, float)):
                    cell.number_format = "#,##0"

    # ---- Column widths ----
    ws.column_dimensions[get_column_letter(LEVEL_COL)].width = 8
    for i in range(len(hid_list)):
        base = start_col + 4 * i
        ws.column_dimensions[get_column_letter(base)].width = 18  # Keep
        ws.column_dimensions[get_column_letter(base + 1)].width = 18  # Success
        ws.column_dimensions[get_column_letter(base + 2)].width = 18  # Break
        ws.column_dimensions[get_column_letter(base + 3)].width = 14  # Samples

    # ---- Row heights ----
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    for lv in range(1, max_level + 1):
        r = row0 + (lv - 1) * 3
        ws.row_dimensions[r].height = 18
        ws.row_dimensions[r + 1].height = 18
        ws.row_dimensions[r + 2].height = 18

    ws.freeze_panes = ws["B3"]

    wb.save(out_xlsx)
